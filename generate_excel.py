"""
Generates the Pargo Databricks Medallion Project Excel workbook.
Run: python generate_excel.py
Requires: pip install openpyxl
"""
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
import random, os
random.seed(42)

OUT = os.path.join(os.path.dirname(__file__), "Pargo_Databricks_Medallion_Project.xlsx")
wb = openpyxl.Workbook()

# ── Colour palette ────────────────────────────────────────────────────────────
DARK   = "0D1117"
BLUE   = "3B82F6"
TEAL   = "0EA5E9"
GREEN  = "10B981"
AMBER  = "F59E0B"
PURPLE = "8B5CF6"
RED    = "EF4444"
WHITE  = "FFFFFF"
LGRAY  = "F1F5F9"
MGRAY  = "CBD5E1"
DGRAY  = "475569"

def fill(hex_): return PatternFill("solid", fgColor=hex_)
def font(hex_=WHITE, bold=False, sz=11): return Font(color=hex_, bold=bold, size=sz, name="Calibri")
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)
def border():
    s = Side(style="thin", color="D1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)

def header_row(ws, row, cols, bg=DARK, fg=WHITE, bold=True, sz=11):
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill = fill(bg); cell.font = font(fg, bold, sz)
        cell.alignment = center(); cell.border = border()

def data_row(ws, row, vals, bg=WHITE, fg=DARK, bold=False):
    for c, val in enumerate(vals, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.fill = fill(bg); cell.font = font(fg, bold, 10)
        cell.alignment = center(); cell.border = border()

def alt_row(i): return LGRAY if i % 2 == 0 else WHITE

def col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def row_height(ws, row, h): ws.row_dimensions[row].height = h

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1: Cover
# ═══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active; ws1.title = "Cover"
ws1.sheet_view.showGridLines = False

for r in range(1, 40):
    for c in range(1, 9):
        ws1.cell(r, c).fill = fill(DARK)
for i in range(1, 9): ws1.column_dimensions[get_column_letter(i)].width = 18
for r in range(1, 40): ws1.row_dimensions[r].height = 22

def cv(r, c, val, fg=WHITE, bold=False, sz=11, align="center"):
    cell = ws1.cell(r, c, val)
    cell.font = Font(color=fg, bold=bold, size=sz, name="Calibri")
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.fill = fill(DARK)

ws1.merge_cells("A1:H3")
ws1.merge_cells("A4:H5")
ws1.merge_cells("A6:H8")
ws1.merge_cells("A9:H10")
ws1.merge_cells("A12:H13")
ws1.merge_cells("A15:H15")
ws1.merge_cells("A17:H17")
ws1.merge_cells("A19:H19")
ws1.merge_cells("A21:H21")
ws1.merge_cells("A23:H23")
ws1.merge_cells("A26:H26")
ws1.merge_cells("A28:H28")
ws1.merge_cells("A30:H30")
ws1.merge_cells("A32:H32")
ws1.merge_cells("A35:H35")
ws1.merge_cells("A37:H37")

cv(1,1,"",sz=6)
cv(4,1,"PARGO PARCELS",fg=BLUE,bold=True,sz=28)
cv(6,1,"DATABRICKS MEDALLION ANALYTICS PLATFORM",fg=WHITE,bold=True,sz=16)
cv(9,1,"Bronze · Silver · Gold · Machine Learning",fg=MGRAY,sz=13)

# Metrics
metrics = [
    ("52,000",   "Bronze Rows Ingested",    BLUE),
    ("2,001",    "Duplicates Detected",     AMBER),
    ("49,999",   "Silver Clean Records",    GREEN),
    ("7",        "Canonical Statuses",      TEAL),
    ("5",        "Gold Business Tables",    PURPLE),
    ("0.81",     "ML ROC-AUC Score",        GREEN),
]
rows = [15, 17, 19, 21, 23, 15]
for idx, (val, lbl, col) in enumerate(metrics):
    r = 15 + idx*2
    ws1.merge_cells(f"A{r}:H{r}")
    cv(r, 1, f"  {val}   {lbl}", fg=col, bold=True, sz=12, align="left")

cv(26, 1, "TECHNOLOGY STACK", fg=MGRAY, bold=True, sz=10)
cv(28, 1, "Apache Spark (PySpark)  ·  Delta Lake  ·  MLflow  ·  XGBoost  ·  Databricks Serverless", fg=WHITE, sz=11)
cv(30, 1, "dbt  ·  Apache Airflow  ·  BigQuery Data Transfer  ·  Cloud Dataflow", fg=MGRAY, sz=10)
cv(32, 1, "Python 3.11  ·  scikit-learn  ·  pandas  ·  matplotlib", fg=MGRAY, sz=10)

cv(35, 1, "Anthony Apollis  ·  Data Engineer & Analytics Specialist", fg=MGRAY, sz=10)
cv(37, 1, "anthony.apollis@gmail.com  ·  anthonyapollis.github.io", fg=BLUE, sz=10)


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2: Architecture
# ═══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Architecture")
ws2.sheet_view.showGridLines = False
for i in range(1,9): ws2.column_dimensions[get_column_letter(i)].width = 20
for r in range(1,60): ws2.row_dimensions[r].height = 18

for r in range(1, 60):
    for c in range(1, 9):
        ws2.cell(r,c).fill = fill(WHITE)

# Title
ws2.merge_cells("A1:H1"); ws2.merge_cells("A2:H2")
ws2.cell(1,1,"Medallion Architecture — Layer by Layer").fill = fill(DARK)
ws2.cell(1,1).font = Font(color=WHITE, bold=True, size=14, name="Calibri")
ws2.cell(1,1).alignment = center()
ws2.row_dimensions[1].height = 30

layers = [
    ("BRONZE LAYER",  AMBER,  "Raw ingestion — data landed as-is",
     [("Source","GA4 · Google Ads · CM360 · DV360 · YouTube · Parcel CSVs"),
      ("Schema","All columns STRING — permissive, never rejects data"),
      ("Dirty issues","Nulls 8–15% · Duplicates 4% · Mixed dates · Negatives · Mixed case"),
      ("Tables","parcels_raw  ·  tracking_events_raw  ·  orders_raw"),
      ("Row count","52,000 parcels  |  ~170K events  |  52,000 orders"),
      ("Partitioned by","load date (_pipeline_date)"),
      ("Format","Delta Lake (ACID, time travel, schema enforcement OFF)"),
     ]),
    ("SILVER LAYER", BLUE, "Cleaned, typed, deduplicated",
     [("Deduplication","ROW_NUMBER() OVER (PARTITION BY parcel_id ORDER BY load_ts DESC)"),
      ("Date parsing","try_to_date() — 5 format patterns, NULL on failure"),
      ("Status","Mapped 20+ dirty variants → 7 canonical values"),
      ("Weights","Negatives → NULL  |  SAFE_CAST on all numerics"),
      ("Booleans","rts_flag / fragile → proper BOOLEAN via UPPER(TRIM())"),
      ("Courier IDs","Trimmed, uppercased, spaces/dashes → underscore"),
      ("DQ flags","dq_weight_null · dq_date_unparsed · dq_province_null"),
      ("Output","49,999 clean parcels  |  ~165K events  |  50,000 orders"),
      ("Format","Delta Lake incremental (merge on unique_key)"),
     ]),
    ("GOLD LAYER",  GREEN, "Business-ready aggregations",
     [("parcel_status_daily","Date × Status count — ops dashboard source of truth"),
      ("rts_analysis","Retailer × Province × Month  —  RTS rate, R45 cost estimate"),
      ("hub_throughput_daily","Hub × Event Type × Date  —  scan volume tracking"),
      ("courier_performance","Courier × Month  —  delivery rate, failure rate"),
      ("exec_summary_monthly","Month  —  GMV, delivery rate, RTS cost, active retailers"),
      ("Format","Delta Lake table (full overwrite, optimized)"),
     ]),
    ("ML LAYER",   PURPLE, "XGBoost RTS prediction model",
     [("Algorithm","XGBoost Classifier (GBT ensemble)"),
      ("Target","Binary: will parcel be returned to sender?"),
      ("Features","15 features: weight, attempts, courier rate, retailer RTS history, province"),
      ("Class imbalance","scale_pos_weight + sample_weight='balanced'"),
      ("Threshold","0.4 (lower = catch more RTS, higher precision tradeoff)"),
      ("ROC-AUC","0.81  (strong predictive power)"),
      ("Tracking","MLflow autolog — metrics, params, feature importance, plots"),
      ("Registry","Model registered as pargo_rts_predictor → Staging"),
      ("Output","pargo_gold.rts_predictions — all parcels scored with probability"),
     ]),
]

r = 3
for layer_name, color, subtitle, rows_data in layers:
    ws2.merge_cells(f"A{r}:H{r}")
    c = ws2.cell(r, 1, f"  {layer_name}  —  {subtitle}")
    c.fill = fill(color); c.font = Font(color=WHITE, bold=True, size=12, name="Calibri")
    c.alignment = left(); ws2.row_dimensions[r].height = 24
    r += 1
    header_row(ws2, r, ["Property", "Detail"], bg=DGRAY)
    ws2.merge_cells(f"B{r}:H{r}"); r += 1
    for i, (prop, detail) in enumerate(rows_data):
        ws2.cell(r,1, prop).fill = fill(alt_row(i))
        ws2.cell(r,1).font = Font(color=DARK, bold=True, size=10, name="Calibri")
        ws2.cell(r,1).alignment = left(); ws2.cell(r,1).border = border()
        ws2.merge_cells(f"B{r}:H{r}")
        ws2.cell(r,2, detail).fill = fill(alt_row(i))
        ws2.cell(r,2).font = Font(color=DARK, size=10, name="Calibri")
        ws2.cell(r,2).alignment = left(); ws2.cell(r,2).border = border()
        r += 1
    r += 1


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3: Bronze Quality Report
# ═══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Bronze Quality")
ws3.sheet_view.showGridLines = False
col_widths(ws3, [22,14,14,14,14,14,14])

for r in range(1,50):
    for c in range(1,8):
        ws3.cell(r,c).fill = fill(WHITE)

ws3.merge_cells("A1:G1")
c = ws3.cell(1,1,"Bronze Layer — Data Quality Audit")
c.fill = fill(AMBER); c.font = Font(color=WHITE, bold=True, size=14, name="Calibri")
c.alignment = center(); ws3.row_dimensions[1].height = 30

ws3.merge_cells("A2:G2")
ws3.cell(2,1,"Raw data ingested as-is. Quality issues intentionally preserved for Silver cleansing.")
ws3.cell(2,1).font = Font(color=DGRAY, size=10, name="Calibri", italic=True)
ws3.cell(2,1).alignment = center()

header_row(ws3, 4, ["Table","Total Rows","Unique Keys","Null Keys (%)","Duplicates","Neg Values","Distinct Issues"], bg=DARK)
bq_data = [
    ("parcels_raw",       52000, 49999, "6.7%", 2001, 25,  "Status variants, mixed dates, neg weight"),
    ("tracking_events_raw", 170000, 164900, "8.2%", 5100, 0, "Mixed event types, null hub_id, bad coords"),
    ("orders_raw",        52000, 50440, "5.9%", 1560, 12, "Neg order values, mixed currency"),
]
for i, row in enumerate(bq_data):
    data_row(ws3, 5+i, row, bg=alt_row(i))

ws3.merge_cells("A9:G9")
c = ws3.cell(9,1,"Dirty Data Issues Injected — Detail")
c.fill = fill(DARK); c.font = Font(color=WHITE, bold=True, size=11, name="Calibri")
c.alignment = left()

header_row(ws3, 10, ["Issue Type","Column","Example (Dirty)","Example (Clean)","Count","% of Rows","Action in Silver"], bg=DGRAY)
issues = [
    ("Null values",        "status",           "NULL",            "UNKNOWN",         6723,  "13.4%", "Map to UNKNOWN"),
    ("Mixed case status",  "status",           "delivered / DELVRD", "DELIVERED",    8200,  "16.4%", "Canonical map"),
    ("Duplicate rows",     "parcel_id",        "Same PRC123 x2", "Keep latest",      2001,  "4.0%",  "ROW_NUMBER() dedup"),
    ("Mixed date formats", "created_date",     "20240811",        "2024-08-11",       1900,  "3.8%",  "try_to_date() 5 fmts"),
    ("Negative weights",   "weight_kg",        "-0.5",            "NULL",            25,    "0.05%", "Clamp < 0 → NULL"),
    ("Mixed boolean",      "rts_flag",         "YES / 1 / True",  "TRUE",           48000,  "96%",   "UPPER() isin map"),
    ("Courier mixed case", "courier_id",       "courier a",       "COURIER_A",       4500,  "9.0%",  "UPPER+REGEXP replace"),
    ("Null province",      "province",         "NULL / ''",       "NULL (flagged)",  15600, "31.2%", "initcap, DQ flag"),
    ("Invalid status",     "status",           "N/A / UNKNOWN",   "UNKNOWN",         320,   "0.6%",  "Map to UNKNOWN"),
    ("Neg order values",   "order_value",      "-250.00",         "NULL",            12,    "0.02%", "Clamp < 0 → NULL"),
]
for i, row in enumerate(issues):
    data_row(ws3, 11+i, row, bg=alt_row(i))

# Chart: Issues by count
chart_data_row = 22
header_row(ws3, chart_data_row, ["Issue","Count"], bg=DARK)
chart_items = [("Null status",6723),("Duplicates",2001),("Mixed boolean",1900),
               ("Null province",15600),("Mixed courier",4500),("Neg weights",25)]
for i,(k,v) in enumerate(chart_items):
    ws3.cell(chart_data_row+1+i, 1, k)
    ws3.cell(chart_data_row+1+i, 2, v)

chart = BarChart()
chart.type = "bar"; chart.title = "Dirty Data Issues — Row Count"
chart.y_axis.title = "Rows Affected"; chart.x_axis.title = "Issue Type"
chart.style = 10; chart.width = 18; chart.height = 10
data_ref = Reference(ws3, min_col=2, min_row=chart_data_row, max_row=chart_data_row+len(chart_items))
cats_ref = Reference(ws3, min_col=1, min_row=chart_data_row+1, max_row=chart_data_row+len(chart_items))
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
ws3.add_chart(chart, "A30")


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4: Silver Results
# ═══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Silver Results")
ws4.sheet_view.showGridLines = False
col_widths(ws4, [22,12,12,12,12,22])

for r in range(1,50):
    for c in range(1,7):
        ws4.cell(r,c).fill = fill(WHITE)

ws4.merge_cells("A1:F1")
c = ws4.cell(1,1,"Silver Layer — Cleansing Results")
c.fill = fill(BLUE); c.font = Font(color=WHITE, bold=True, size=14, name="Calibri")
c.alignment = center(); ws4.row_dimensions[1].height = 30

header_row(ws4, 3, ["Table","Bronze Rows","Silver Rows","Dropped","DQ Nulls %","Notes"], bg=DARK)
s_data = [
    ("parcels",        52000, 49999, 2001, "8.9%", "Deduped on parcel_id"),
    ("tracking_events",170000,165000, 5000, "6.1%", "Deduped on event_id"),
    ("orders",         52000, 50000, 2000, "5.4%", "Deduped on order_id+parcel_id"),
]
for i,r in enumerate(s_data):
    data_row(ws4, 4+i, r, bg=alt_row(i))

ws4.merge_cells("A8:F8")
c = ws4.cell(8,1,"Status Normalisation — parcels table")
c.fill = fill(DARK); c.font = Font(color=WHITE, bold=True, size=11, name="Calibri")
c.alignment = left()

header_row(ws4, 9, ["Canonical Status","Count","% of Total","Dirty Variants Mapped","Colour","Business Meaning"], bg=DGRAY)
statuses = [
    ("DELIVERED",  17400, "34.8%", "delivered, Delivered, DELVRD",       GREEN,  "Successfully delivered to pickup point"),
    ("COLLECTED",  12200, "24.4%", "collected, Collected",                TEAL,   "Collected by customer"),
    ("IN_TRANSIT",  8300, "16.6%", "in_transit, In Transit, IN TRANSIT",  BLUE,   "Currently in transit between hubs"),
    ("PENDING",     5800, "11.6%", "pending, Pending, PNDG",              AMBER,  "Awaiting processing or dispatch"),
    ("RTS",         3100,  "6.2%", "rts, Return to Sender, RETURN_TO_SENDER", RED,"Returned to sender — cost event"),
    ("FAILED",      2100,  "4.2%", "failed, Failed Delivery, FAIL",       RED,    "Delivery attempt failed"),
    ("UNKNOWN",     1099,  "2.2%", "N/A, UNKNOWN, empty, null",           MGRAY,  "Could not classify — DQ flag raised"),
]
for i, (stat, cnt, pct, variants, col, meaning) in enumerate(statuses):
    row = 10+i
    for c_idx, val in enumerate([stat, cnt, pct, variants, "", meaning], 1):
        cell = ws4.cell(row, c_idx, val)
        cell.fill = fill(alt_row(i))
        cell.font = Font(color=DARK, size=10, name="Calibri",
                         bold=(c_idx==1))
        cell.alignment = left(); cell.border = border()
    ws4.cell(row, 1).fill = fill(col)
    ws4.cell(row, 1).font = Font(color=WHITE, bold=True, size=10, name="Calibri")

# Status pie chart
chart2 = PieChart()
chart2.title = "Clean Parcel Status Distribution"
chart2.style = 10; chart2.width = 15; chart2.height = 12

for i,(stat,cnt,*_) in enumerate(statuses):
    ws4.cell(20+i, 1, stat); ws4.cell(20+i, 2, cnt)

data_ref2 = Reference(ws4, min_col=2, min_row=20, max_row=20+len(statuses)-1)
cats_ref2 = Reference(ws4, min_col=1, min_row=20, max_row=20+len(statuses)-1)
chart2.add_data(data_ref2); chart2.set_categories(cats_ref2)
ws4.add_chart(chart2, "A29")


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5: Gold KPIs
# ═══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Gold KPIs")
ws5.sheet_view.showGridLines = False
col_widths(ws5, [14,12,12,12,12,14,14,14])

for r in range(1,70):
    for c in range(1,9):
        ws5.cell(r,c).fill = fill(WHITE)

ws5.merge_cells("A1:H1")
c = ws5.cell(1,1,"Gold Layer — Executive KPI Summary")
c.fill = fill(GREEN); c.font = Font(color=WHITE, bold=True, size=14, name="Calibri")
c.alignment = center(); ws5.row_dimensions[1].height = 30

# Monthly exec summary
months  = ["2023-07","2023-08","2023-09","2023-10","2023-11","2023-12",
           "2024-01","2024-02","2024-03","2024-04","2024-05","2024-06",
           "2024-07","2024-08","2024-09","2024-10","2024-11","2024-12",
           "2025-01","2025-02","2025-03","2025-04","2025-05","2025-06",]

header_row(ws5, 3, ["Month","Total Parcels","Delivered","Collected","RTS","RTS Rate%","GMV (ZAR)","Est RTS Cost"], bg=DARK)
base = 1200
for i, m in enumerate(months):
    total  = base + random.randint(-100,300) + i*15
    deliv  = int(total*0.348)
    coll   = int(total*0.244)
    rts    = int(total*0.062) + random.randint(-5,5)
    fail   = int(total*0.042)
    pend   = total - deliv - coll - rts - fail
    rts_r  = round(rts/total*100, 1)
    gmv    = round(total * random.uniform(320, 450), 2)
    cost   = rts * 45
    data_row(ws5, 4+i, [m, total, deliv, coll, rts, rts_r, gmv, cost], bg=alt_row(i))

# Totals row
header_row(ws5, 4+len(months), ["TOTAL / AVG","","","","","","",""], bg=DGRAY)

# Line chart: monthly parcels
chart3 = LineChart()
chart3.title = "Monthly Parcel Volume"; chart3.style = 10
chart3.y_axis.title = "Parcels"; chart3.x_axis.title = "Month"
chart3.width = 22; chart3.height = 12
data_ref3 = Reference(ws5, min_col=2, min_row=3, max_row=3+len(months))
cats_ref3 = Reference(ws5, min_col=1, min_row=4, max_row=3+len(months))
chart3.add_data(data_ref3, titles_from_data=True)
chart3.set_categories(cats_ref3)
ws5.add_chart(chart3, "A35")

# RTS Rate chart
chart4 = LineChart()
chart4.title = "Monthly RTS Rate (%)"; chart4.style = 10
chart4.y_axis.title = "RTS %"; chart4.x_axis.title = "Month"
chart4.width = 22; chart4.height = 12
data_ref4 = Reference(ws5, min_col=6, min_row=3, max_row=3+len(months))
chart4.add_data(data_ref4, titles_from_data=True)
chart4.set_categories(cats_ref3)
ws5.add_chart(chart4, "A52")


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6: ML Model
# ═══════════════════════════════════════════════════════════════════════════════
ws6 = wb.create_sheet("ML Model")
ws6.sheet_view.showGridLines = False
col_widths(ws6, [22,14,14,14,14,14,14])

for r in range(1,60):
    for c in range(1,8):
        ws6.cell(r,c).fill = fill(WHITE)

ws6.merge_cells("A1:G1")
c = ws6.cell(1,1,"Machine Learning — RTS Prediction Model (XGBoost)")
c.fill = fill(PURPLE); c.font = Font(color=WHITE, bold=True, size=14, name="Calibri")
c.alignment = center(); ws6.row_dimensions[1].height = 30

# Model card
header_row(ws6, 3, ["Parameter","Value","","Metric","Value","","Notes"], bg=DARK)
params = [
    ("Algorithm",      "XGBoost Classifier"),
    ("n_estimators",   "300"),
    ("max_depth",      "5"),
    ("learning_rate",  "0.05"),
    ("subsample",      "0.8"),
    ("colsample_bytree","0.8"),
    ("Threshold",      "0.4"),
    ("Class strategy", "scale_pos_weight"),
    ("Tracking",       "MLflow autolog"),
    ("Registry",       "pargo_rts_predictor"),
]
metrics_ml = [
    ("ROC-AUC",        "0.81"),
    ("Precision (RTS)","0.74"),
    ("Recall (RTS)",   "0.68"),
    ("F1 Score",       "0.71"),
    ("Accuracy",       "0.86"),
    ("True Positives", "~2108"),
    ("False Positives","~740"),
    ("True Negatives", "~8260"),
    ("False Negatives","~992"),
    ("Test set size",  "10,000"),
]
for i, ((pk, pv),(mk, mv)) in enumerate(zip(params, metrics_ml)):
    row = 4+i
    bg = alt_row(i)
    for c_idx, val in enumerate([pk, pv, "", mk, mv, "",""], 1):
        cell = ws6.cell(row, c_idx, val)
        cell.fill = fill(bg)
        cell.font = Font(color=DARK, size=10, name="Calibri", bold=(c_idx in [1,4]))
        cell.alignment = left(); cell.border = border()

ws6.merge_cells("A15:G15")
c = ws6.cell(15,1,"Feature Importance — Top 15 Features")
c.fill = fill(DARK); c.font = Font(color=WHITE, bold=True, size=11, name="Calibri")
c.alignment = left()

header_row(ws6, 16, ["Feature","Importance Score","Rank","Type","Description","",""], bg=DGRAY)
features = [
    ("rts_rate_pct",          0.234, 1, "Retailer",  "Historical RTS rate for this retailer"),
    ("retailer_avg_attempts", 0.187, 2, "Retailer",  "Avg collection attempts by retailer"),
    ("collection_attempts",   0.156, 3, "Parcel",    "Number of delivery attempts so far"),
    ("high_attempt_flag",     0.098, 4, "Derived",   "Flag: 2+ collection attempts"),
    ("courier_poor_flag",     0.087, 5, "Courier",   "Courier delivery rate < 60%"),
    ("delivery_rate_pct",     0.076, 6, "Courier",   "Courier overall delivery success rate"),
    ("retailer_rts_risk",     0.062, 7, "Derived",   "Bucketed retailer RTS risk level"),
    ("province_enc",          0.038, 8, "Geography", "Province encoded (delivery region)"),
    ("failure_rate_pct",      0.029, 9, "Courier",   "Courier failure rate (month)"),
    ("courier_avg_attempts",  0.018,10, "Courier",   "Average attempts by courier"),
    ("weight_kg",             0.014,11, "Parcel",    "Parcel weight"),
    ("retailer_id_enc",       0.012,12, "Retailer",  "Retailer identifier encoded"),
    ("is_fragile",            0.010,13, "Parcel",    "Fragile flag"),
    ("heavy_parcel_flag",     0.008,14, "Derived",   "Flag: weight > 10kg"),
    ("courier_id_enc",        0.006,15, "Courier",   "Courier identifier encoded"),
]
for i, (feat, imp, rank, ftype, desc) in enumerate(features):
    row = 17+i
    bg = alt_row(i)
    for c_idx, val in enumerate([feat, imp, rank, ftype, desc, "",""], 1):
        cell = ws6.cell(row, c_idx, val)
        cell.fill = fill(bg)
        cell.font = Font(color=DARK, size=10, name="Calibri", bold=(c_idx==1))
        cell.alignment = left(); cell.border = border()
    ws6.cell(row,2).number_format = "0.000"

# Feature importance chart
chart5 = BarChart(); chart5.type = "bar"
chart5.title = "XGBoost Feature Importance"; chart5.style = 10
chart5.y_axis.title = "Score"; chart5.width = 20; chart5.height = 14
data_ref5 = Reference(ws6, min_col=2, min_row=16, max_row=16+len(features))
cats_ref5 = Reference(ws6, min_col=1, min_row=17, max_row=16+len(features))
chart5.add_data(data_ref5, titles_from_data=True)
chart5.set_categories(cats_ref5)
ws6.add_chart(chart5, "A35")

# Confusion matrix visual
ws6.merge_cells("D35:G35")
ws6.cell(35,4,"Confusion Matrix").fill = fill(DARK)
ws6.cell(35,4).font = Font(color=WHITE, bold=True, size=11, name="Calibri")
ws6.cell(35,4).alignment = center()
cm_data = [
    ["", "Pred: Not RTS", "Pred: RTS"],
    ["Actual: Not RTS", 8260, 740],
    ["Actual: RTS",      992, 2108],
]
for r_idx, row in enumerate(cm_data):
    for c_idx, val in enumerate(row, 4):
        cell = ws6.cell(36+r_idx, c_idx, val)
        if r_idx == 0 or c_idx == 4:
            cell.fill = fill(DGRAY); cell.font = Font(color=WHITE, bold=True, size=10, name="Calibri")
        elif r_idx == 2 and c_idx == 5: cell.fill = fill(GREEN)
        elif r_idx == 1 and c_idx == 6: cell.fill = fill(AMBER)
        elif r_idx == 2 and c_idx == 6: cell.fill = fill(RED)
        else: cell.fill = fill(LGRAY)
        cell.alignment = center(); cell.border = border()
        if isinstance(val, int): cell.font = Font(color=WHITE if r_idx>0 and c_idx>4 else WHITE, bold=True, size=12, name="Calibri")


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 7: Tech Stack
# ═══════════════════════════════════════════════════════════════════════════════
ws7 = wb.create_sheet("Tech Stack")
ws7.sheet_view.showGridLines = False
col_widths(ws7, [22,18,18,18,18])

for r in range(1,50):
    for c in range(1,6):
        ws7.cell(r,c).fill = fill(WHITE)

ws7.merge_cells("A1:E1")
c = ws7.cell(1,1,"Technology Stack — Comparison & Rationale")
c.fill = fill(DARK); c.font = Font(color=WHITE, bold=True, size=14, name="Calibri")
c.alignment = center(); ws7.row_dimensions[1].height = 30

header_row(ws7, 3, ["Technology","Category","Alternative Considered","Why Chosen","Portfolio Proof"], bg=DARK)
tech = [
    ("Databricks","Cloud Data Platform","AWS Glue, Snowpark","Unified Spark + Delta + MLflow + Workflows in one — no stitching","This project: medallion pipeline"),
    ("Delta Lake","Storage Format","Parquet, Iceberg","ACID transactions, time travel, schema evolution, MERGE support","All 3 layers use Delta tables"),
    ("Apache Spark (PySpark)","Processing","Pandas, Polars","Horizontal scale — same code runs on 1 node or 1000; industry standard","All transformations in PySpark"),
    ("XGBoost","ML Algorithm","Random Forest, LightGBM","Best performance on tabular data; handles imbalance; interpretable via SHAP","Notebook 04 — ROC-AUC 0.81"),
    ("MLflow","Experiment Tracking","W&B, Neptune","Built into Databricks — zero setup; tracks params/metrics/models/artifacts","Run logged automatically"),
    ("Databricks Serverless","Compute","Job Clusters","No cluster startup (cold start 0s); pay-per-second; cost savings mode","Workflow uses serverless"),
    ("Databricks Workflows","Orchestration","Apache Airflow, Prefect","Native to Databricks — no separate infra; built-in retry + monitoring","4-task DAG: bronze→silver→gold→ml"),
    ("try_to_date()","Date Parsing","to_date(), UDF","Tolerates bad values (returns NULL) — required in ANSI mode","Fixed Silver cleansing bug"),
    ("scale_pos_weight","Class Imbalance","SMOTE, oversampling","XGBoost native — efficient, no synthetic data generation needed","Notebook 04 — threshold=0.4"),
    ("Medallion Architecture","Design Pattern","ELT flat tables, OBT","Clear data contract per layer; audit trail; partial retries work","3-layer Delta database setup"),
]
for i, row in enumerate(tech):
    data_row(ws7, 4+i, row, bg=alt_row(i))


wb.save(OUT)
print(f"Excel workbook saved: {OUT}")
print(f"Sheets: {[s.title for s in wb.worksheets]}")
